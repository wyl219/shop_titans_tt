# import pandas as pd
# from en2zh import en2zh
#
# def edit_head(cell_content):
#     # 这里可以添加对表头内容的处理逻辑
#     return cell_content
#
# # 读取xlsx文件
# file_path = "1.xlsx"
# output_path = "./2.xlsx"
# df = pd.read_excel(file_path)
#
# # 在最左侧加一列，表头为 "名称"，其余行与第二列内容相同
# df.insert(1, '名称', df.iloc[:, 0])
#
# # 处理表头
# df.columns = [edit_head(col) for col in df.columns]
#
# # 处理其余行，除第一个单元格外
# for index, row in df.iterrows():
#     for col in df.columns[1:]:
#         if isinstance(row[col], str):
#             df.at[index, col] = en2zh(row[col])
# # 另存为 "./2.xlsx"
#
# df.to_excel(output_path, index=False)
#
# print(f"文件已保存到 {output_path}")
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from com.en2zh import en2zh

def edit_head(cell_content):
    # 这里可以添加对表头内容的处理逻辑
    return cell_content

# 读取xlsx文件
file_path = "../1.xlsx"
output_path = "./2.xlsx"
df = pd.read_excel(file_path)

# 在最左侧加一列，表头为 "名称"，其余行与第二列内容相同
df.insert(1, '名称', df.iloc[:, 0])

# 处理表头
df.columns = [edit_head(col) for col in df.columns]

# 处理其余行，除第一个单元格外
for index, row in df.iterrows():
    for col in df.columns[1:]:
        if isinstance(row[col], str):
            df.at[index, col] = en2zh(row[col])

# 使用openpyxl加载原始Excel文件
book = load_workbook(file_path)
sheet = book.active  # 选择活动工作表

# 添加新的列名到工作表的第一行
for col_index, column_name in enumerate(df.columns, start=1):
    sheet.cell(row=1, column=col_index).value = column_name

# 将DataFrame的每一行添加到工作表中，从第二行开始
for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
    for col_index, value in enumerate(row, start=1):
        sheet.cell(row=row_index, column=col_index).value = value

# 保存工作簿
book.save(output_path)
print(f"文件已保存到 {output_path}")
